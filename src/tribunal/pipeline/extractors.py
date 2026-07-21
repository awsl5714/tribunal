"""Evidence extraction from source documents (PDF / DOCX / XLSX).

Thin, dependency-optional wrappers. In the reference project these fed hundreds
of pages of PDF and Word plus Excel review sheets into the scoring pipeline. The
extractors return plain text keyed by a locator so downstream evidence keeps a
precise citation.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path


@dataclass
class ExtractedPage:
    locator: str
    text: str


def extract_pdf(path: str | Path) -> list[ExtractedPage]:
    """Extract text per page. Requires ``pypdf`` (optional dependency)."""
    from pypdf import PdfReader  # lazy import

    reader = PdfReader(str(path))
    pages: list[ExtractedPage] = []
    for i, page in enumerate(reader.pages, start=1):
        pages.append(ExtractedPage(locator=f"p.{i}", text=page.extract_text() or ""))
    return pages


def extract_docx(path: str | Path) -> list[ExtractedPage]:
    """Extract paragraphs. Requires ``python-docx`` (optional dependency)."""
    import docx  # lazy import

    document = docx.Document(str(path))
    return [
        ExtractedPage(locator=f"para.{i}", text=p.text)
        for i, p in enumerate(document.paragraphs, start=1)
        if p.text.strip()
    ]


def extract_xlsx(path: str | Path, sheet: str | None = None) -> list[ExtractedPage]:
    """Extract non-empty cells. Requires ``openpyxl`` (optional dependency)."""
    from openpyxl import load_workbook  # lazy import

    wb = load_workbook(str(path), data_only=True)
    ws = wb[sheet] if sheet else wb.active
    pages: list[ExtractedPage] = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                pages.append(
                    ExtractedPage(locator=f"{ws.title}!{cell.coordinate}", text=str(cell.value))
                )
    return pages


def extract(path: str | Path) -> list[ExtractedPage]:
    """Dispatch on file extension."""
    suffix = Path(path).suffix.lower()
    if suffix == ".pdf":
        return extract_pdf(path)
    if suffix in (".docx", ".doc"):
        return extract_docx(path)
    if suffix in (".xlsx", ".xlsm"):
        return extract_xlsx(path)
    raise ValueError(f"unsupported source type: {suffix}")
